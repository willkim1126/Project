import pandas as pd
from datetime import datetime, time
from typing import List, Tuple
from dataclasses import dataclass, field
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


@dataclass
class Employee:
    employee_id: str
    name: str
    position: str
    location_preferences: List[str]
    preferred_hours: dict  # Dictionary with days as keys and time ranges as values
    current_assignments: dict = field(default_factory=dict)

    def __post_init__(self):
        self.current_assignments = {day: [] for day in self.preferred_hours.keys()}

    def is_available(self, day: str, start_time: time, end_time: time) -> bool:
        """Check if employee is available for a given time slot."""
        for assigned_start, assigned_end, _ in self.current_assignments.get(day, []):
            if start_time < assigned_end and end_time > assigned_start:
                return False
        return True

    def add_assignment(self, day: str, start_time: time, end_time: time, location: str) -> None:
        """Add a new assignment to the employee's schedule."""
        if day not in self.current_assignments:
            self.current_assignments[day] = []
        self.current_assignments[day].append((start_time, end_time, location))
        self.current_assignments[day].sort(key=lambda x: x[0])


class ScheduleOptimizer:
    def __init__(self):
        self.days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        self.locations = ['CSC', 'The Trove', 'Seventh', 'ERC']  # Prioritize CSC first
        self.time_frames = {
            location: (time(7, 0), time(15, 30)) for location in self.locations  # Ensure shifts end by 3:30 PM
        }
        self.employees = []
        self.csc_assignment_count = 0  # Counter for CSC assignments (limit to 10)

    def load_employees_from_file(self, file_path: str) -> None:
        """Load employee data from an Excel file."""
        df = pd.read_excel(file_path)
        
        for _, row in df.iterrows():
            preferred_hours = {
                day: self._parse_time_ranges(row[day]) if pd.notna(row[day]) else []
                for day in self.days
            }
            employee = Employee(
                employee_id=row['employee_id'],
                name=row['name'],
                position=row['position'],
                location_preferences=[loc.strip() for loc in row['location_preferences'].split(',')],
                preferred_hours=preferred_hours,
            )
            self.employees.append(employee)

    def _parse_time_ranges(self, time_str: str) -> List[Tuple[time, time]]:
        """Parse time ranges from a string."""
        ranges = []
        pairs = time_str.split(',')
        for pair in pairs:
            start_str, end_str = pair.strip().split('-')
            start_time = datetime.strptime(start_str.strip(), '%H:%M').time()
            end_time = datetime.strptime(end_str.strip(), '%H:%M').time()
            if end_time <= time(15, 30):  # Ensure shifts end by 3:30 PM
                ranges.append((start_time, end_time))
        return ranges

    def optimize_schedule(self) -> None:
        """Optimize the schedule by assigning employees to shifts."""
        max_hours_per_week = 19.5

        for employee in sorted(self.employees, key=lambda x: x.position.lower() == "lead"):
            total_hours_assigned = 0

            for day in self.days:
                if total_hours_assigned >= max_hours_per_week:
                    break

                for start_time, end_time in employee.preferred_hours.get(day, []):
                    shift_duration = (datetime.combine(datetime.today(), end_time) -
                                      datetime.combine(datetime.today(), start_time)).seconds / 3600

                    if shift_duration < 2 or total_hours_assigned + shift_duration > max_hours_per_week:
                        continue

                    # Assign shift prioritizing CSC first (limit to 10 assignments)
                    for location in sorted(employee.location_preferences,
                                           key=lambda loc: loc == "CSC", reverse=True):
                        if location == "CSC" and self.csc_assignment_count >= 10:
                            continue

                        loc_start, loc_end = self.time_frames[location]
                        if start_time < loc_start or end_time > loc_end:
                            continue

                        if employee.is_available(day, start_time, end_time):
                            employee.add_assignment(day, start_time, end_time, location)
                            total_hours_assigned += shift_duration

                            # Increment CSC assignment count if assigned to CSC
                            if location == "CSC":
                                self.csc_assignment_count += 1

                            break

    def export_schedule_to_excel(self, output_file: str) -> None:
        """Export the optimized schedule to an Excel file."""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Optimized Schedule"

        # Define styles
        header_font = Font(bold=True)
        location_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

        # Write headers (days)
        col_idx = 1
        for day in self.days:
            sheet.cell(row=1, column=col_idx).value = day
            sheet.cell(row=1, column=col_idx).font = header_font
            col_idx += 1

        # Write schedule under each day column
        for col_idx, day in enumerate(self.days, start=1):
            row_idx = 2
            for location in self.locations:
                # Write location name with formatting
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = location.upper()
                cell.font = header_font
                cell.fill = location_fill
                row_idx += 1

                # Write employees assigned to this location on this day
                for employee in self.employees:
                    assignments = employee.current_assignments.get(day)
                    if assignments:
                        for start_time, end_time, loc in assignments:
                            if loc == location:
                                sheet.cell(row=row_idx, column=col_idx).value = (
                                    f"{employee.name} ({start_time.strftime('%H:%M')}-{end_time.strftime('%H:%M')})"
                                )
                                row_idx += 1

        workbook.save(output_file)
        print(f"\nSchedule exported to {output_file}")


if __name__ == '__main__':
    optimizer = ScheduleOptimizer()
    optimizer.load_employees_from_file('example_split_shifts.xlsx')
    optimizer.optimize_schedule()
    optimizer.export_schedule_to_excel('optimized_schedule.xlsx')
